from pathlib import Path
import csv
import json
import re
import numpy as np
import cv2
import rawpy
import imageio.v3 as iio
from skimage.color import rgb2lab, deltaE_ciede2000
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# CONFIG
# ============================================================

PROJECT_DIR = Path(".")
BEFORE_DIR = PROJECT_DIR / "before"
AFTER_DIR = PROJECT_DIR / "after"
OUTPUT_DIR = PROJECT_DIR / "outputs"
DEBUG_DIR = OUTPUT_DIR / "debug"

OUTPUT_DIR.mkdir(exist_ok=True, parents=True)
DEBUG_DIR.mkdir(exist_ok=True, parents=True)

RAW_EXTENSIONS = {".orf", ".ORF", ".raw", ".RAW"}
STD_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".tif", ".tiff"}
VALID_EXTENSIONS = RAW_EXTENSIONS | STD_IMAGE_EXTENSIONS

# If using flat folders, group is inferred from the first character.
# Edit this if your naming convention differs.
GROUP_PREFIX_MAP = {
    "C": "control",
    "S": "treated",
    "T": "treated",
}

# Segmentation settings
SEGMENTATION_MODE = "auto"   # "auto" or "center_crop"
CENTER_CROP_FRACTION = 0.82
ERODE_BORDER_FRACTION = 0.03
MIN_MASK_FILL_FRACTION = 0.05
MAX_MASK_FILL_FRACTION = 0.95

# Residual dirty threshold settings
# Learned from each tile's BEFORE image using a/b chromatic variation only
# MIN_AB_THRESHOLD = 2.0
MIN_AB_THRESHOLD = 2.0
BASELINE_PERCENTILE = 95
AB_MARGIN = 1.5

# Debug outputs
SAVE_DEBUG_MASKS = True
SAVE_DEBUG_CROPS = True
SAVE_OVERLAYS = True

# ============================================================
# SORTING / NAMING HELPERS
# ============================================================

def natural_tile_sort_key(tile_id: str):
    """
    Natural sort for IDs like C1, C2, C10, S1, S12, etc.
    """
    parts = re.split(r"(\d+)", str(tile_id))
    key = []
    for part in parts:
        if part.isdigit():
            key.append(int(part))
        else:
            key.append(part.upper())
    return key

def pairing_key(path: Path) -> str:
    """
    Normalize a filename stem for matching before/after pairs.
    Examples:
      T1_before.ORF -> t1
      C2-after.ORF  -> c2
      S3.ORF        -> s3
    """
    s = path.stem.strip().lower()
    for suffix in [
        "_before", "-before", " before",
        "_after", "-after", " after",
        "_pre", "-pre", " pre",
        "_post", "-post", " post",
    ]:
        if s.endswith(suffix):
            s = s[: -len(suffix)].strip()
            break
    return s

def infer_group_from_key(key: str) -> str:
    if not key:
        raise ValueError("Empty pairing key")
    prefix = key[0].upper()
    if prefix not in GROUP_PREFIX_MAP:
        raise ValueError(
            f"Could not infer group from '{key}'. "
            f"Edit GROUP_PREFIX_MAP at the top of the script."
        )
    return GROUP_PREFIX_MAP[prefix]

# ============================================================
# DISCOVERY
# ============================================================

def list_image_files(folder: Path):
    return sorted(
        [p for p in folder.iterdir() if p.is_file() and p.suffix in VALID_EXTENSIONS]
    )

def discover_tiles_subfolders(before_dir: Path, after_dir: Path):
    """
    Supports:
      before/treated/*.ORF
      before/control/*.ORF
      after/treated/*.ORF
      after/control/*.ORF
    """
    groups = ["treated", "control"]
    tiles = []

    for group in groups:
        before_group = before_dir / group
        after_group = after_dir / group

        if not before_group.exists() or not after_group.exists():
            return None

        before_files = {pairing_key(p): p for p in list_image_files(before_group)}
        after_files = {pairing_key(p): p for p in list_image_files(after_group)}

        shared = sorted(set(before_files) & set(after_files))
        if not shared:
            print(f"Warning: no matching files found for subfolder group '{group}'")
            continue

        missing_before = sorted(set(after_files) - set(before_files))
        missing_after = sorted(set(before_files) - set(after_files))

        if missing_before:
            print(f"Warning: {group} AFTER files with no BEFORE match: {missing_before}")
        if missing_after:
            print(f"Warning: {group} BEFORE files with no AFTER match: {missing_after}")

        for key in shared:
            tiles.append({
                "tile_id": key.upper(),
                "group": group,
                "before_path": before_files[key],
                "after_path": after_files[key],
            })

    return tiles

def discover_tiles_flat(before_dir: Path, after_dir: Path):
    """
    Supports:
      before/C1.ORF, before/S1.ORF, ...
      after/C1.ORF, after/S1.ORF, ...
    """
    before_files = {pairing_key(p): p for p in list_image_files(before_dir)}
    after_files = {pairing_key(p): p for p in list_image_files(after_dir)}

    shared = sorted(set(before_files) & set(after_files))
    if not shared:
        raise ValueError("No matching image pairs found between before/ and after/")

    missing_before = sorted(set(after_files) - set(before_files))
    missing_after = sorted(set(before_files) - set(after_files))

    if missing_before:
        print(f"Warning: AFTER files with no BEFORE match: {missing_before}")
    if missing_after:
        print(f"Warning: BEFORE files with no AFTER match: {missing_after}")

    tiles = []
    for key in shared:
        tiles.append({
            "tile_id": key.upper(),
            "group": infer_group_from_key(key),
            "before_path": before_files[key],
            "after_path": after_files[key],
        })

    return tiles

def discover_tiles(before_dir: Path, after_dir: Path):
    subfolder_tiles = discover_tiles_subfolders(before_dir, after_dir)
    if subfolder_tiles:
        print("Using treated/control subfolder layout.")
        return subfolder_tiles

    print("Using flat-folder layout.")
    return discover_tiles_flat(before_dir, after_dir)

# ============================================================
# IMAGE IO
# ============================================================

def linear_to_srgb(img: np.ndarray) -> np.ndarray:
    img = np.clip(img, 0.0, 1.0)
    return np.where(
        img <= 0.0031308,
        12.92 * img,
        1.055 * np.power(img, 1 / 2.4) - 0.055
    ).clip(0.0, 1.0)

def read_image_rgb(path: Path) -> np.ndarray:
    """
    Returns RGB float32 image in [0,1].
    RAW files are standardized with no auto-brightness.
    """
    if path.suffix in RAW_EXTENSIONS:
        with rawpy.imread(str(path)) as raw:
            rgb16 = raw.postprocess(
                gamma=(1, 1),
                no_auto_bright=True,
                output_bps=16
            )
        rgb = rgb16.astype(np.float32) / 65535.0
        return linear_to_srgb(rgb)

    if path.suffix.lower() in STD_IMAGE_EXTENSIONS:
        arr = iio.imread(path)
        if arr.ndim == 2:
            arr = np.stack([arr, arr, arr], axis=-1)
        if arr.shape[-1] == 4:
            arr = arr[..., :3]

        if arr.dtype == np.uint8:
            return arr.astype(np.float32) / 255.0
        if arr.dtype == np.uint16:
            return arr.astype(np.float32) / 65535.0

        arr = arr.astype(np.float32)
        arr /= max(arr.max(), 1.0)
        return np.clip(arr, 0.0, 1.0)

    raise ValueError(f"Unsupported image extension: {path.suffix}")

# ============================================================
# SEGMENTATION / CROPPING
# ============================================================

def center_crop_mask(h: int, w: int, frac: float):
    mask = np.zeros((h, w), dtype=np.uint8)
    ch = int(round(h * frac))
    cw = int(round(w * frac))
    y0 = max(0, (h - ch) // 2)
    x0 = max(0, (w - cw) // 2)
    mask[y0:y0 + ch, x0:x0 + cw] = 255
    return mask

def largest_reasonable_mask(gray_u8: np.ndarray):
    blur = cv2.GaussianBlur(gray_u8, (7, 7), 0)

    _, th = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    _, th_inv = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    h, w = gray_u8.shape
    img_area = h * w
    candidates = []

    for candidate in (th, th_inv):
        candidate = cv2.morphologyEx(candidate, cv2.MORPH_CLOSE, np.ones((9, 9), np.uint8))
        candidate = cv2.morphologyEx(candidate, cv2.MORPH_OPEN, np.ones((5, 5), np.uint8))
        contours, _ = cv2.findContours(candidate, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

        for c in contours:
            area = cv2.contourArea(c)
            if area < 0.03 * img_area:
                continue

            mask = np.zeros_like(gray_u8, dtype=np.uint8)
            cv2.drawContours(mask, [c], -1, 255, thickness=cv2.FILLED)
            fill_frac = np.mean(mask > 0)

            if not (MIN_MASK_FILL_FRACTION <= fill_frac <= MAX_MASK_FILL_FRACTION):
                continue

            x, y, bw, bh = cv2.boundingRect(c)
            rect_area = max(bw * bh, 1)
            rectangularity = area / rect_area

            candidates.append((area, rectangularity, mask))

    if not candidates:
        return np.zeros_like(gray_u8, dtype=np.uint8)

    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    # Pick whichever candidate contains the center pixel —
    # tile is always centered enough that the center belongs to it
    cy, cx = gray_u8.shape[0] // 2, gray_u8.shape[1] // 2
    for _, _, mask in candidates:
        if mask[cy, cx] > 0:
            return mask
    return candidates[0][2]   # fallback if nothing contains center
    # return candidates[0][2]

def erode_mask(mask: np.ndarray, frac: float):
    ys, xs = np.where(mask > 0)
    if len(xs) == 0 or len(ys) == 0:
        return mask

    h = ys.max() - ys.min() + 1
    w = xs.max() - xs.min() + 1
    k = max(3, int(round(min(h, w) * frac)))
    if k % 2 == 0:
        k += 1

    kernel = np.ones((k, k), np.uint8)
    return cv2.erode(mask, kernel, iterations=1)

def detect_and_crop_tile(rgb: np.ndarray):
    h, w = rgb.shape[:2]
    cy, cx = h // 2, w // 2

    if SEGMENTATION_MODE == "center_crop":
        mask = center_crop_mask(h, w, CENTER_CROP_FRACTION)
    else:
        gray = cv2.cvtColor((rgb * 255).astype(np.uint8), cv2.COLOR_RGB2GRAY)
        mask = largest_reasonable_mask(gray)
        if mask.sum() == 0 or mask[cy, cx] == 0:
            print("Warning: auto segmentation missed center pixel, falling back to center crop.")
            # print("Warning: auto segmentation failed, falling back to center crop.")
            mask = center_crop_mask(h, w, CENTER_CROP_FRACTION)

    mask = erode_mask(mask, ERODE_BORDER_FRACTION)

    ys, xs = np.where(mask > 0)
    if len(xs) == 0 or len(ys) == 0:
        raise ValueError("Tile mask is empty after erosion")

    y0, y1 = ys.min(), ys.max()
    x0, x1 = xs.min(), xs.max()

    crop = rgb[y0:y1 + 1, x0:x1 + 1]
    crop_mask = mask[y0:y1 + 1, x0:x1 + 1]

    return crop, crop_mask

# ============================================================
# METRICS
# ============================================================

def extract_lab(rgb: np.ndarray):
    return rgb2lab(rgb)

def tile_pixels_from_mask(lab: np.ndarray, mask: np.ndarray):
    pixels = lab[mask > 0]
    if len(pixels) < 500:
        raise ValueError("Too few tile pixels after masking")
    return pixels

def ciede2000_triplets(lab1: np.ndarray, lab2: np.ndarray) -> float:
    arr1 = np.asarray(lab1, dtype=np.float32).reshape(1, 1, 3)
    arr2 = np.asarray(lab2, dtype=np.float32).reshape(1, 1, 3)
    return float(deltaE_ciede2000(arr1, arr2)[0, 0])

def ab_distance(lab1: np.ndarray, lab2: np.ndarray) -> float:
    return float(np.linalg.norm(np.asarray(lab1)[1:3] - np.asarray(lab2)[1:3]))

def build_before_baseline(before_pixels: np.ndarray):
    """
    Learn the tile's normal chromatic variation from the BEFORE image.
    Dirty threshold is based on a/b only, which is less sensitive to lighting.
    """
    before_median = np.median(before_pixels, axis=0)

    baseline_ab_residuals = np.linalg.norm(
        before_pixels[:, 1:3] - before_median[1:3],
        axis=1
    )

    threshold_ab = max(
        MIN_AB_THRESHOLD,
        float(np.percentile(baseline_ab_residuals, BASELINE_PERCENTILE)) + AB_MARGIN
    )

    return {
        "before_median_lab": before_median,
        "before_threshold_ab": threshold_ab,
        "before_baseline_ab_p95": float(np.percentile(baseline_ab_residuals, 95)),
        "before_baseline_ab_mean": float(np.mean(baseline_ab_residuals)),
    }

def score_after_against_before(
    before_pixels: np.ndarray,
    after_pixels: np.ndarray,
    after_lab_img: np.ndarray,
    after_mask: np.ndarray,
    threshold_ab: float
):
    before_median = np.median(before_pixels, axis=0)
    after_median = np.median(after_pixels, axis=0)

    # Overall tile shift between sessions
    shift_lab = after_median - before_median

    # Correct the AFTER image by removing the global tile shift
    corrected_after_lab = after_lab_img.copy()
    corrected_after_lab[..., 0] -= shift_lab[0]
    corrected_after_lab[..., 1] -= shift_lab[1]
    corrected_after_lab[..., 2] -= shift_lab[2]

    corrected_after_pixels = corrected_after_lab[after_mask > 0]

    # Residual dirtyness after removing the global shift, using a/b only
    residual_ab = np.linalg.norm(
        corrected_after_pixels[:, 1:3] - before_median[1:3],
        axis=1
    )

    # dirty_vector = residual_ab > threshold_ab
    # Only flag pixels shifting toward green (negative a*) — ignores shadow artifacts
    a_shift = corrected_after_pixels[:, 1] - before_median[1]
    dirty_vector = (residual_ab > threshold_ab) & (a_shift < 0)
    residual_dirty_percent_ab = 100.0 * float(np.mean(dirty_vector))
    cleanliness_score = 100.0 - residual_dirty_percent_ab

    # Tile-level median change metrics
    median_deltaE00_full = ciede2000_triplets(before_median, after_median)
    median_delta_ab = ab_distance(before_median, after_median)
    median_delta_L = float(after_median[0] - before_median[0])

    return {
        "before_median_lab": before_median,
        "after_median_lab": after_median,
        "shift_lab": shift_lab,
        "threshold_ab": float(threshold_ab),
        "median_deltaE00_full": round(median_deltaE00_full, 2),
        "median_delta_ab": round(median_delta_ab, 2),
        "median_delta_L": round(median_delta_L, 2),
        "residual_dirty_percent_ab": round(residual_dirty_percent_ab, 2),
        "cleanliness_score": round(cleanliness_score, 2),
        "residual_ab_mean": round(float(np.mean(residual_ab)), 2),
        "residual_ab_p90": round(float(np.percentile(residual_ab, 90)), 2),
        "dirty_vector": dirty_vector,
    }

# ============================================================
# OUTPUT HELPERS
# ============================================================

def save_csv(path: Path, rows):
    if not rows:
        return
    with open(path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)

def save_mask_preview(rgb: np.ndarray, mask: np.ndarray, out_path: Path):
    preview = (rgb * 255).astype(np.uint8).copy()
    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cv2.drawContours(preview, contours, -1, (0, 255, 0), 2)
    iio.imwrite(out_path, preview)

def save_overlay(rgb: np.ndarray, mask: np.ndarray, dirty_vector: np.ndarray, out_path: Path):
    img = (rgb * 255).astype(np.uint8).copy()
    dirty_mask = np.zeros(mask.shape, dtype=bool)
    dirty_mask[mask > 0] = dirty_vector

    overlay = img.copy()
    overlay[dirty_mask] = [255, 0, 0]

    alpha = 0.45
    blended = img.copy()
    blended[dirty_mask] = (
        (1 - alpha) * img[dirty_mask] + alpha * overlay[dirty_mask]
    ).astype(np.uint8)

    contours, _ = cv2.findContours(mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    cv2.drawContours(blended, contours, -1, (0, 255, 0), 2)
    iio.imwrite(out_path, blended)

def mean_or_blank(values):
    return round(float(np.mean(values)), 2) if values else ""

def save_human_readable_excel(path: Path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tile Summary"

    sorted_rows = sorted(rows, key=lambda r: natural_tile_sort_key(r["tile_id"]))

    headers = [
        "Tile",
        "Group",
        "Median ΔE00",
        "Median Δab",
        "Dirty %",
    ]
    ws.append(headers)

    # Styling
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    summary_fill = PatternFill(fill_type="solid", fgColor="EDEDED")

    header_font = Font(name="Avenir", bold=True, size=11)
    body_font = Font(name="Avenir", size=11)
    summary_font = Font(name="Avenir", bold=True, size=11)

    center = Alignment(horizontal="center", vertical="center")

    thin_side = Side(style="thin", color="BFBFBF")
    all_borders = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side,
    )

    # Header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = all_borders

    # Data rows
    for r in sorted_rows:
        group_label = "Control" if r["group"] == "control" else "Treated"
        ws.append([
            r["tile_id"],
            group_label,
            r["median_deltaE00_full"],
            r["median_delta_ab"],
            r["residual_dirty_percent_ab"],
        ])

    # Apply body styling + borders to data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.font = body_font
            cell.alignment = center
            cell.border = all_borders

    # Blank row before summaries
    ws.append([])
    blank_row_idx = ws.max_row
    for cell in ws[blank_row_idx]:
        cell.border = all_borders
        cell.font = body_font
        cell.alignment = center

    control_rows = [r for r in sorted_rows if r["group"] == "control"]
    treated_rows = [r for r in sorted_rows if r["group"] == "treated"]

    control_summary = [
        "Control Average",
        "Control",
        mean_or_blank([r["median_deltaE00_full"] for r in control_rows]),
        mean_or_blank([r["median_delta_ab"] for r in control_rows]),
        mean_or_blank([r["residual_dirty_percent_ab"] for r in control_rows]),
    ]

    treated_summary = [
        "Treated Average",
        "Treated",
        mean_or_blank([r["median_deltaE00_full"] for r in treated_rows]),
        mean_or_blank([r["median_delta_ab"] for r in treated_rows]),
        mean_or_blank([r["residual_dirty_percent_ab"] for r in treated_rows]),
    ]

    ws.append(control_summary)
    ws.append(treated_summary)

    # Style summary rows
    for row_idx in [ws.max_row - 1, ws.max_row]:
        for cell in ws[row_idx]:
            cell.font = summary_font
            cell.fill = summary_fill
            cell.alignment = center
            cell.border = all_borders

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12

    # Format numeric columns
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=5):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    ws.freeze_panes = "A2"
    wb.save(path)

# ============================================================
# MAIN
# ============================================================

def main():
    if not BEFORE_DIR.exists():
        raise FileNotFoundError(f"Missing folder: {BEFORE_DIR}")
    if not AFTER_DIR.exists():
        raise FileNotFoundError(f"Missing folder: {AFTER_DIR}")

    tiles = discover_tiles(BEFORE_DIR, AFTER_DIR)
    if not tiles:
        raise ValueError("No matched tile pairs found.")

    tiles = sorted(tiles, key=lambda t: natural_tile_sort_key(t["tile_id"]))

    print("\nDiscovered tiles:")
    for t in tiles:
        print(
            f"  {t['tile_id']} | {t['group']} | "
            f"{t['before_path'].name} -> {t['after_path'].name}"
        )

    per_tile_rows = []

    print("\nProcessing tiles...\n")

    for tile in tiles:
        tile_id = tile["tile_id"]
        group = tile["group"]
        before_path = tile["before_path"]
        after_path = tile["after_path"]

        before_rgb = read_image_rgb(before_path)
        after_rgb = read_image_rgb(after_path)

        before_crop, before_mask = detect_and_crop_tile(before_rgb)
        after_crop, after_mask = detect_and_crop_tile(after_rgb)

        before_lab_img = extract_lab(before_crop)
        after_lab_img = extract_lab(after_crop)

        before_pixels = tile_pixels_from_mask(before_lab_img, before_mask)
        after_pixels = tile_pixels_from_mask(after_lab_img, after_mask)

        baseline = build_before_baseline(before_pixels)
        result = score_after_against_before(
            before_pixels=before_pixels,
            after_pixels=after_pixels,
            after_lab_img=after_lab_img,
            after_mask=after_mask,
            threshold_ab=baseline["before_threshold_ab"],
        )

        row = {
            "tile_id": tile_id,
            "group": group,
            "before_file": before_path.name,
            "after_file": after_path.name,

            "before_median_L": round(float(result["before_median_lab"][0]), 4),
            "before_median_a": round(float(result["before_median_lab"][1]), 4),
            "before_median_b": round(float(result["before_median_lab"][2]), 4),

            "after_median_L": round(float(result["after_median_lab"][0]), 4),
            "after_median_a": round(float(result["after_median_lab"][1]), 4),
            "after_median_b": round(float(result["after_median_lab"][2]), 4),

            "global_shift_L": round(float(result["shift_lab"][0]), 4),
            "global_shift_a": round(float(result["shift_lab"][1]), 4),
            "global_shift_b": round(float(result["shift_lab"][2]), 4),

            "baseline_threshold_ab": round(float(baseline["before_threshold_ab"]), 4),
            "baseline_ab_p95": round(float(baseline["before_baseline_ab_p95"]), 4),
            "baseline_ab_mean": round(float(baseline["before_baseline_ab_mean"]), 4),

            "median_deltaE00_full": result["median_deltaE00_full"],
            "median_delta_ab": result["median_delta_ab"],
            "median_delta_L": result["median_delta_L"],

            "residual_dirty_percent_ab": result["residual_dirty_percent_ab"],
            "cleanliness_score": result["cleanliness_score"],
            "residual_ab_mean": result["residual_ab_mean"],
            "residual_ab_p90": result["residual_ab_p90"],
        }
        per_tile_rows.append(row)

        if SAVE_DEBUG_CROPS:
            iio.imwrite(DEBUG_DIR / f"{tile_id}_before_crop.png", (before_crop * 255).astype(np.uint8))
            iio.imwrite(DEBUG_DIR / f"{tile_id}_after_crop.png", (after_crop * 255).astype(np.uint8))

        if SAVE_DEBUG_MASKS:
            save_mask_preview(before_crop, before_mask, DEBUG_DIR / f"{tile_id}_before_mask.png")
            save_mask_preview(after_crop, after_mask, DEBUG_DIR / f"{tile_id}_after_mask.png")

        if SAVE_OVERLAYS:
            save_overlay(after_crop, after_mask, result["dirty_vector"], DEBUG_DIR / f"{tile_id}_residual_overlay.png")

        print(
            f"{tile_id}: "
            f"median ΔE00={result['median_deltaE00_full']}, "
            f"median Δab={result['median_delta_ab']}, "
            f"dirty={result['residual_dirty_percent_ab']}%"
        )

    per_tile_rows = sorted(per_tile_rows, key=lambda r: natural_tile_sort_key(r["tile_id"]))

    # Group summaries
    group_summary_rows = []
    groups = ["control", "treated"]

    for group in groups:
        rows = [r for r in per_tile_rows if r["group"] == group]
        if not rows:
            continue

        de_vals = np.array([r["median_deltaE00_full"] for r in rows], dtype=float)
        dab_vals = np.array([r["median_delta_ab"] for r in rows], dtype=float)
        dirty_vals = np.array([r["residual_dirty_percent_ab"] for r in rows], dtype=float)
        clean_vals = np.array([r["cleanliness_score"] for r in rows], dtype=float)

        group_summary_rows.append({
            "group": group,
            "n_tiles": len(rows),
            "mean_median_deltaE00_full": round(float(np.mean(de_vals)), 2),
            "std_median_deltaE00_full": round(float(np.std(de_vals, ddof=0)), 2),
            "mean_median_delta_ab": round(float(np.mean(dab_vals)), 2),
            "std_median_delta_ab": round(float(np.std(dab_vals, ddof=0)), 2),
            "mean_residual_dirty_percent_ab": round(float(np.mean(dirty_vals)), 2),
            "std_residual_dirty_percent_ab": round(float(np.std(dirty_vals, ddof=0)), 2),
            "mean_cleanliness_score": round(float(np.mean(clean_vals)), 2),
            "std_cleanliness_score": round(float(np.std(clean_vals, ddof=0)), 2),
        })

    # Save outputs
    save_csv(OUTPUT_DIR / "per_tile_results.csv", per_tile_rows)
    save_csv(OUTPUT_DIR / "group_summary.csv", group_summary_rows)
    save_human_readable_excel(OUTPUT_DIR / "tile_summary.xlsx", per_tile_rows)

    summary_json = {
        "tiles": per_tile_rows,
        "group_summary": group_summary_rows,
    }
    with open(OUTPUT_DIR / "summary.json", "w") as f:
        json.dump(summary_json, f, indent=2)

    print("\nDone.")
    print(f"Saved: {OUTPUT_DIR / 'per_tile_results.csv'}")
    print(f"Saved: {OUTPUT_DIR / 'group_summary.csv'}")
    print(f"Saved: {OUTPUT_DIR / 'tile_summary.xlsx'}")
    print(f"Saved: {OUTPUT_DIR / 'summary.json'}")
    print(f"Debug images: {DEBUG_DIR}")

if __name__ == "__main__":
    main()